"""
ElevateIQ — Export Service
Generates CSV and XLSX result exports for admin download.
"""
import csv
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from flask import make_response

from sqlalchemy import func
from models.models import Submission, db


# ─────────────────────────────────────────────
# Shared data fetcher
# ─────────────────────────────────────────────

def _get_results(search: str = '', assessment_id: int = None, status: str = None):
    """Fetch all submissions with optional search and status filter."""
    from sqlalchemy.orm import joinedload
    query = (
        db.session.query(Submission)
        .options(
            joinedload(Submission.candidate),
            joinedload(Submission.assessment)
        )
        .filter(Submission.status != 'in_progress')
        .order_by(
            Submission.percentage.desc(),
            Submission.score.desc(),
            Submission.submitted_at.desc()
        )
    )

    if status and status.lower() in ('pass', 'fail'):
        query = query.filter(Submission.status == status.lower())

    if assessment_id:
        query = query.filter(Submission.assessment_id == assessment_id)

    if search:
        from models.models import Candidate
        like = f'%{search}%'
        query = query.join(Submission.candidate).filter(
            db.or_(
                Candidate.full_name.ilike(like),
                Candidate.hall_ticket.ilike(like),
                Candidate.email.ilike(like),
            )
        )

    return query.all()


HEADERS = [
    'Hall Ticket', 'Full Name', 'Email',
    'Assessment', 'Score', 'Total Questions',
    'Percentage (%)', 'Violations', 'Status', 'Submitted At'
]


def _row_data(sub: Submission):
    return [
        sub.candidate.hall_ticket,
        sub.candidate.full_name,
        sub.candidate.email,
        sub.assessment.title,
        sub.score,
        sub.total_questions,
        round(sub.percentage, 2),
        sub.violations,
        sub.status.upper(),
        sub.submitted_at_ist.strftime('%Y-%m-%d %H:%M:%S') if sub.submitted_at_ist else '',
    ]


# ─────────────────────────────────────────────
# CSV Export
# ─────────────────────────────────────────────

def export_csv(search: str = '', assessment_id: int = None, status: str = None):
    """Return a Flask streaming Response with CSV attachment."""
    from flask import Response

    def generate():
        # Write header
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(HEADERS)
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)

        # Stream rows
        submissions = _get_results(search, assessment_id, status)
        for sub in submissions:
            writer.writerow(_row_data(sub))
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

    status_suffix = f'_{status.lower()}' if status and status.lower() in ('pass', 'fail') else ''
    filename = f'elevateiq_results{status_suffix}_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.csv'
    return Response(
        generate(),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────
# XLSX Export
# ─────────────────────────────────────────────

def export_xlsx(search: str = '', assessment_id: int = None, status: str = None):
    """Return a Flask Response with XLSX attachment."""
    submissions = _get_results(search, assessment_id, status)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Results'

    # ── Styles ──────────────────────────────
    header_fill = PatternFill('solid', fgColor='2D1B69')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    pass_fill = PatternFill('solid', fgColor='D1FAE5')
    fail_fill = PatternFill('solid', fgColor='FEE2E2')
    alt_fill = PatternFill('solid', fgColor='F5F3FF')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='C4B5FD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = [18, 28, 32, 30, 8, 16, 16, 12, 10, 22]

    # ── Title row ──────────────────────────
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    status_title = f' ({status.upper()} ONLY)' if status and status.lower() in ('pass', 'fail') else ''
    title_cell.value = f'ElevateIQ — Assessment Results Export{status_title}  ({datetime.utcnow().strftime("%d %b %Y %H:%M UTC")})'
    title_cell.font = Font(bold=True, color='4C1D95', size=13)
    title_cell.fill = PatternFill('solid', fgColor='EDE9FE')
    title_cell.alignment = center
    ws.row_dimensions[1].height = 26

    # ── Header row ─────────────────────────
    ws.append(HEADERS)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 22

    # ── Data rows ──────────────────────────
    for row_idx, sub in enumerate(submissions, start=3):
        row = _row_data(sub)
        ws.append(row)
        status_upper = sub.status.upper()
        row_fill = pass_fill if status_upper == 'PASS' else (
            fail_fill if status_upper == 'FAIL' else None
        )
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            elif row_idx % 2 == 0:
                cell.fill = alt_fill
            if col_idx in (1, 5, 6, 7, 8, 9):
                cell.alignment = center
            else:
                cell.alignment = left

    # ── Column widths ──────────────────────
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Freeze header rows ─────────────────
    ws.freeze_panes = 'A3'

    # ── Auto filter ────────────────────────
    if submissions:
        ws.auto_filter.ref = f'A2:J{len(submissions) + 2}'

    # ── Stream response ────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    status_suffix = f'_{status.lower()}' if status and status.lower() in ('pass', 'fail') else ''
    filename = f'elevateiq_results{status_suffix}_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.headers['Content-Type'] = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    return response


# ─────────────────────────────────────────────
# Daily Assessment Reports Helper & Exports
# ─────────────────────────────────────────────

def _get_daily_report_data(start_date: str = '', end_date: str = '', assessment_id: int = None):
    """Fetch daily aggregated assessment data."""
    date_col = func.date(Submission.submitted_at)

    query = (
        db.session.query(
            date_col.label('report_date'),
            func.count(Submission.id).label('total'),
            func.sum(db.case((Submission.status == 'pass', 1), else_=0)).label('passed'),
            func.sum(db.case((Submission.status == 'fail', 1), else_=0)).label('failed'),
            func.avg(Submission.percentage).label('avg_percentage'),
            func.max(Submission.score).label('max_score'),
            func.min(Submission.score).label('min_score')
        )
        .filter(Submission.status != 'in_progress', Submission.submitted_at.isnot(None))
    )

    if assessment_id:
        query = query.filter(Submission.assessment_id == assessment_id)
    if start_date:
        query = query.filter(date_col >= start_date)
    if end_date:
        query = query.filter(date_col <= end_date)

    daily_stats = (
        query.group_by(date_col)
        .order_by(date_col.desc())
        .all()
    )

    reports = []
    for row in daily_stats:
        t = row.total or 0
        p = int(row.passed or 0)
        f = int(row.failed or 0)
        pass_rate = round((p / t * 100), 1) if t > 0 else 0.0
        avg_pct = round(row.avg_percentage or 0.0, 1)

        reports.append({
            'date': str(row.report_date),
            'total': t,
            'passed': p,
            'failed': f,
            'pass_rate': pass_rate,
            'avg_percentage': avg_pct,
            'max_score': row.max_score or 0,
            'min_score': row.min_score or 0
        })
    return reports


DAILY_HEADERS = [
    'Date', 'Total Candidates', 'Passed Count', 'Failed Count',
    'Pass Rate (%)', 'Avg Score (%)', 'Max Score', 'Min Score'
]


def export_daily_reports_csv(start_date: str = '', end_date: str = '', assessment_id: int = None):
    """Return CSV download of daily assessment reports."""
    from flask import Response
    reports = _get_daily_report_data(start_date, end_date, assessment_id)

    def generate():
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(DAILY_HEADERS)
        yield output.getvalue()
        output.seek(0)
        output.truncate(0)

        for r in reports:
            writer.writerow([
                r['date'], r['total'], r['passed'], r['failed'],
                r['pass_rate'], r['avg_percentage'], r['max_score'], r['min_score']
            ])
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

    filename = f'elevateiq_daily_assessment_reports_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.csv'
    return Response(
        generate(),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


def export_daily_reports_xlsx(start_date: str = '', end_date: str = '', assessment_id: int = None):
    """Return XLSX download of daily assessment reports."""
    reports = _get_daily_report_data(start_date, end_date, assessment_id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Daily Reports'

    header_fill = PatternFill('solid', fgColor='2D1B69')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    pass_fill = PatternFill('solid', fgColor='D1FAE5')
    alt_fill = PatternFill('solid', fgColor='F5F3FF')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='C4B5FD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    col_widths = [16, 18, 16, 16, 16, 16, 14, 14]

    # Title row
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f'ElevateIQ — Day-by-Day Assessment Reports Export ({datetime.utcnow().strftime("%d %b %Y %H:%M UTC")})'
    title_cell.font = Font(bold=True, color='4C1D95', size=13)
    title_cell.fill = PatternFill('solid', fgColor='EDE9FE')
    title_cell.alignment = center
    ws.row_dimensions[1].height = 26

    # Header row
    ws.append(DAILY_HEADERS)
    for col_idx, header in enumerate(DAILY_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 22

    # Data rows
    for row_idx, r in enumerate(reports, start=3):
        row_values = [
            r['date'], r['total'], r['passed'], r['failed'],
            r['pass_rate'], r['avg_percentage'], r['max_score'], r['min_score']
        ]
        ws.append(row_values)
        for col_idx in range(1, len(DAILY_HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = center

    # Column widths & freeze
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = 'A3'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'elevateiq_daily_assessment_reports_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = make_response(output.getvalue())
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response

